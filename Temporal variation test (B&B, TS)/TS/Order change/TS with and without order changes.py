import random
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Set random seed for reproducibility
random.seed(42)

# Define the recipes with their eligibility
recipes_f1_only = list(range(1, 30))
recipes_f1_f2 = list(range(30, 50))
recipes_f2_only = list(range(50, 90))
recipes_f3_only = list(range(90, 101))

# Define total orders and factory capacities
total_orders = 10000
F1_cap = int(0.25 * total_orders)  # 25% of total orders
F2_cap = int(0.5 * total_orders)  # 50% of total orders
factory_capacities = {
    'F1': F1_cap,
    'F2': F2_cap,
    'F3': float('inf')  # F3 has unlimited capacity
}

def generate_order_recipes(eligible_recipes, max_recipes=4):
    return random.sample(eligible_recipes, random.randint(1, min(max_recipes, len(eligible_recipes))))

def generate_orders_for_day(real_proportion, simulated_proportion, existing_real_orders=None, with_changes=True):
    orders = []
    f1_f3_target = int(0.3 * total_orders)
    f2_f3_target = int(0.6 * total_orders)
    f1_f2_f3_target = int(0.1 * total_orders)
    
    total_real_orders = int(real_proportion * total_orders)
    
    changed_orders = []
    deleted_orders = []
    
    if existing_real_orders and with_changes:
        # Delete 5% of real orders
        delete_count = int(0.05 * len(existing_real_orders))
        deleted_orders = random.sample(existing_real_orders, delete_count)
        remaining_orders = [order for order in existing_real_orders if order not in deleted_orders]
        
        # Change 30% of remaining real orders
        change_count = int(0.3 * len(remaining_orders))
        changed_orders = random.sample(remaining_orders, change_count)
        
        for order in remaining_orders:
            if order in changed_orders:
                # Change recipe composition and eligibility
                new_recipes = generate_order_recipes(recipes_f1_only + recipes_f1_f2 + recipes_f2_only + recipes_f3_only)
                new_eligibility = ['F3']  # Always eligible for F3
                if any(r in recipes_f1_only + recipes_f1_f2 for r in new_recipes):
                    new_eligibility.append('F1')
                if any(r in recipes_f1_f2 + recipes_f2_only for r in new_recipes):
                    new_eligibility.append('F2')
                order['recipe_ids'] = new_recipes
                order['eligible_factories'] = new_eligibility
            orders.append(order)
    elif existing_real_orders and not with_changes:
        # Carry over all existing real orders without changes
        orders.extend(existing_real_orders)
    
    # Generate new orders to reach the total_orders
    while len(orders) < total_orders:
        if len([o for o in orders if set(o['eligible_factories']) == {'F1', 'F2', 'F3'}]) < f1_f2_f3_target:
            recipe_ids = generate_order_recipes(recipes_f1_f2)
            eligible_factories = ['F1', 'F2', 'F3']
        elif len([o for o in orders if set(o['eligible_factories']) == {'F1', 'F3'}]) < f1_f3_target - f1_f2_f3_target:
            recipe_ids = generate_order_recipes(recipes_f1_only)
            eligible_factories = ['F1', 'F3']
        elif len([o for o in orders if set(o['eligible_factories']) == {'F2', 'F3'}]) < f2_f3_target - f1_f2_f3_target:
            recipe_ids = generate_order_recipes(recipes_f2_only)
            eligible_factories = ['F2', 'F3']
        else:
            f3_recipe = random.choice(recipes_f3_only)
            all_recipes = recipes_f1_only + recipes_f1_f2 + recipes_f2_only + recipes_f3_only
            remaining_recipes = random.sample(all_recipes, min(3, random.randint(0, 3)))
            recipe_ids = [f3_recipe] + remaining_recipes
            random.shuffle(recipe_ids)
            eligible_factories = ['F3']
        
        new_order = {
            'recipe_ids': recipe_ids,
            'eligible_factories': eligible_factories,
            'is_real': len([o for o in orders if o.get('is_real', False)]) < total_real_orders
        }
        orders.append(new_order)
    
    # Ensure we have the correct number of real orders
    real_order_count = sum(1 for order in orders if order.get('is_real', False))
    if real_order_count < total_real_orders:
        for order in random.sample([o for o in orders if not o.get('is_real', False)], total_real_orders - real_order_count):
            order['is_real'] = True
    elif real_order_count > total_real_orders:
        for order in random.sample([o for o in orders if o.get('is_real', False)], real_order_count - total_real_orders):
            order['is_real'] = False
    
    random.shuffle(orders)
    
    # Assign IDs to orders
    max_id = max([order['id'] for order in existing_real_orders]) if existing_real_orders else 0
    for order in orders:
        if 'id' not in order:
            max_id += 1
            order['id'] = max_id
    
    return orders, changed_orders, deleted_orders

def get_recipe_counts(allocation, by_factory=False):
    if by_factory:
        recipe_counts = {factory: {} for factory in allocation}
        for factory, orders in allocation.items():
            for order in orders:
                for recipe_id in order['recipe_ids']:
                    recipe_counts[factory][recipe_id] = recipe_counts[factory].get(recipe_id, 0) + 1
    else:
        recipe_counts = {}
        for factory, orders in allocation.items():
            for order in orders:
                for recipe_id in order['recipe_ids']:
                    recipe_counts[recipe_id] = recipe_counts.get(recipe_id, 0) + 1
    return recipe_counts

def calculate_total_abs_diff(recipe_counts_t_minus_1, recipe_counts_t):
    all_recipes = set(recipe_counts_t_minus_1['F1'].keys()) | set(recipe_counts_t_minus_1['F2'].keys()) | set(recipe_counts_t_minus_1['F3'].keys()) | \
                  set(recipe_counts_t['F1'].keys()) | set(recipe_counts_t['F2'].keys()) | set(recipe_counts_t['F3'].keys())
    
    total_abs_diff = 0
    
    for recipe_id in all_recipes:
        for factory in ['F1', 'F2', 'F3']:
            t_minus_1_count = recipe_counts_t_minus_1[factory].get(recipe_id, 0)
            t_count = recipe_counts_t[factory].get(recipe_id, 0)
            total_abs_diff += abs(t_count - t_minus_1_count)
    
    return total_abs_diff

def tabu_search(allocation_t_minus_1, allocation_t, factory_capacities, max_iterations=500):
    current_allocation = {factory: orders[:] for factory, orders in allocation_t.items()}
    recipe_counts_t_minus_1 = get_recipe_counts(allocation_t_minus_1, by_factory=True)
    current_recipe_counts = get_recipe_counts(current_allocation, by_factory=True)
    current_total_abs_diff = calculate_total_abs_diff(recipe_counts_t_minus_1, current_recipe_counts)
    
    best_allocation = current_allocation.copy()
    best_total_abs_diff = current_total_abs_diff
    
    tabu_list = {}
    tabu_tenure = 20
    
    def swap_orders(factory1, order1, factory2, order2):
        current_allocation[factory1].remove(order1)
        current_allocation[factory2].remove(order2)
        current_allocation[factory1].append(order2)
        current_allocation[factory2].append(order1)
        
        for recipe_id in order1['recipe_ids']:
            current_recipe_counts[factory1][recipe_id] -= 1
            current_recipe_counts[factory2][recipe_id] = current_recipe_counts[factory2].get(recipe_id, 0) + 1
        for recipe_id in order2['recipe_ids']:
            current_recipe_counts[factory2][recipe_id] -= 1
            current_recipe_counts[factory1][recipe_id] = current_recipe_counts[factory1].get(recipe_id, 0) + 1
    
    def calculate_move_impact(order1, factory1, order2, factory2):
        diff = 0
        for recipe_id in order1['recipe_ids']:
            diff -= abs(current_recipe_counts[factory1][recipe_id] - recipe_counts_t_minus_1[factory1].get(recipe_id, 0))
            diff -= abs(current_recipe_counts[factory2].get(recipe_id, 0) - recipe_counts_t_minus_1[factory2].get(recipe_id, 0))
            diff += abs(current_recipe_counts[factory1][recipe_id] - 1 - recipe_counts_t_minus_1[factory1].get(recipe_id, 0))
            diff += abs(current_recipe_counts[factory2].get(recipe_id, 0) + 1 - recipe_counts_t_minus_1[factory2].get(recipe_id, 0))
        for recipe_id in order2['recipe_ids']:
            diff -= abs(current_recipe_counts[factory2][recipe_id] - recipe_counts_t_minus_1[factory2].get(recipe_id, 0))
            diff -= abs(current_recipe_counts[factory1].get(recipe_id, 0) - recipe_counts_t_minus_1[factory1].get(recipe_id, 0))
            diff += abs(current_recipe_counts[factory2][recipe_id] - 1 - recipe_counts_t_minus_1[factory2].get(recipe_id, 0))
            diff += abs(current_recipe_counts[factory1].get(recipe_id, 0) + 1 - recipe_counts_t_minus_1[factory1].get(recipe_id, 0))
        return diff
    
    for iteration in range(max_iterations):       
        best_move = None
        best_move_diff = 0
        
        # Randomly select a subset of orders to consider for swapping
        orders_to_consider = random.sample(sum(current_allocation.values(), []), min(100, len(sum(current_allocation.values(), []))))
        
        for order1 in orders_to_consider:
            factory1 = next(f for f, orders in current_allocation.items() if order1 in orders)
            for factory2 in factory_capacities:
                if factory1 != factory2:
                    for order2 in random.sample(current_allocation[factory2], min(10, len(current_allocation[factory2]))):
                        if factory2 in order1['eligible_factories'] and factory1 in order2['eligible_factories']:
                            move = (order1['id'], factory1, order2['id'], factory2)
                            if move not in tabu_list or current_total_abs_diff + calculate_move_impact(order1, factory1, order2, factory2) < best_total_abs_diff:
                                diff = calculate_move_impact(order1, factory1, order2, factory2)
                                if diff < best_move_diff:
                                    best_move = (order1, factory1, order2, factory2)
                                    best_move_diff = diff
        
        if best_move:
            order1, factory1, order2, factory2 = best_move
            swap_orders(factory1, order1, factory2, order2)
            current_total_abs_diff += best_move_diff
            
            if current_total_abs_diff < best_total_abs_diff:
                best_total_abs_diff = current_total_abs_diff
                best_allocation = {f: orders[:] for f, orders in current_allocation.items()}
            
            tabu_list[(order1['id'], factory1, order2['id'], factory2)] = iteration + tabu_tenure
            tabu_list[(order2['id'], factory2, order1['id'], factory1)] = iteration + tabu_tenure
        
        # Remove expired tabu moves
        tabu_list = {k: v for k, v in tabu_list.items() if v > iteration}
        
        # Diversification strategy
        if iteration % 100 == 0:
            factory1, factory2 = random.sample(list(factory_capacities.keys()), 2)
            order1 = random.choice(current_allocation[factory1])
            order2 = random.choice(current_allocation[factory2])
            if factory2 in order1['eligible_factories'] and factory1 in order2['eligible_factories']:
                swap_orders(factory1, order1, factory2, order2)
                current_total_abs_diff += calculate_move_impact(order1, factory1, order2, factory2)
    
    total_items_t = sum(sum(counts.values()) for counts in get_recipe_counts(best_allocation, by_factory=True).values())
    best_wmape_site = best_total_abs_diff / total_items_t if total_items_t > 0 else float('inf')

    return best_allocation, best_wmape_site

def allocate_orders(orders, factory_capacities):
    allocation = {factory: [] for factory in factory_capacities}
    remaining_orders = orders.copy()
    f1_eligible = sorted([order for order in remaining_orders if 'F1' in order['eligible_factories']],
                         key=lambda x: len(x['eligible_factories']))
    allocation['F1'] = [order for order in f1_eligible[:factory_capacities['F1']]]
    remaining_orders = [order for order in remaining_orders if order not in allocation['F1']]
    f2_eligible = [order for order in remaining_orders if 'F2' in order['eligible_factories']]
    allocation['F2'] = [order for order in f2_eligible[:factory_capacities['F2']]]
    remaining_orders = [order for order in remaining_orders if order not in allocation['F2']]
    allocation['F3'] = remaining_orders
    return allocation

def calculate_wmape_site(allocation_t_minus_1, allocation_t):
    recipe_counts_t_minus_1 = get_recipe_counts(allocation_t_minus_1, by_factory=True)
    recipe_counts_t = get_recipe_counts(allocation_t, by_factory=True)
    
    total_abs_diff = calculate_total_abs_diff(recipe_counts_t_minus_1, recipe_counts_t)
    total_items_t = sum(sum(counts.values()) for counts in recipe_counts_t.values())
    
    if total_items_t == 0:
        return float('inf')
    else:
        return total_abs_diff / total_items_t

def calculate_wmape_global(allocation_t_minus_1, allocation_t):
    recipe_counts_t_minus_1 = get_recipe_counts(allocation_t_minus_1)
    recipe_counts_t = get_recipe_counts(allocation_t)
    
    all_recipes = set(recipe_counts_t_minus_1.keys()) | set(recipe_counts_t.keys())
    
    total_abs_diff = sum(abs(recipe_counts_t.get(recipe_id, 0) - recipe_counts_t_minus_1.get(recipe_id, 0)) for recipe_id in all_recipes)
    total_items_t = sum(recipe_counts_t.values())
    
    if total_items_t == 0:
        return float('inf')
    else:
        return total_abs_diff / total_items_t

def run_allocation_process_over_time(start_day, end_day, total_orders, with_changes=True):
    allocations_tabu = {}
    wmape_site_values_tabu = []
    wmape_global_values_tabu = []
    real_orders_proportions = []
    previous_real_orders = None
    changed_orders_log = {}
    deleted_orders_log = {}
    
    for day in range(start_day, end_day + 1):
        real_proportion = min(1.0, max(0.1, 0.1 + (0.9 / 15) * (18 + day)))
        simulated_proportion = 1 - real_proportion
        
        orders, changed_orders, deleted_orders = generate_orders_for_day(real_proportion, simulated_proportion, previous_real_orders, with_changes)
        
        if with_changes:
            changed_orders_log[day] = changed_orders
            deleted_orders_log[day] = deleted_orders
        
        allocation_initial = allocate_orders(orders, factory_capacities)
        
        if day == start_day:
            allocation_tabu, _ = tabu_search({f: [] for f in factory_capacities}, allocation_initial, factory_capacities)
            allocations_tabu[day] = allocation_tabu
        else:
            allocation_tabu, wmape_site_tabu = tabu_search(allocations_tabu[day-1], allocation_initial, factory_capacities)
            wmape_site_values_tabu.append(wmape_site_tabu)
            wmape_global_tabu = calculate_wmape_global(allocations_tabu[day-1], allocation_tabu)
            wmape_global_values_tabu.append(wmape_global_tabu)
        
        allocations_tabu[day] = allocation_tabu
        real_orders_proportions.append(len([o for o in orders if o['is_real']]) / total_orders)
        previous_real_orders = [order for order in orders if order['is_real']]
    
    return wmape_site_values_tabu, wmape_global_values_tabu, real_orders_proportions, changed_orders_log, deleted_orders_log

def plot_wmape_comparison(days, wmape_site_with_changes, wmape_site_without_changes, wmape_global_with_changes, wmape_global_without_changes):
    plt.figure(figsize=(14, 8))
    
    plt.plot(days[1:], wmape_site_with_changes, color='tab:red', marker='o', linestyle='-', label='WMAPE site (With changes)')
    plt.plot(days[1:], wmape_global_with_changes, color='tab:green', marker='^', linestyle='--', label='WMAPE global (With changes)')
    plt.plot(days[1:], wmape_site_without_changes, color='tab:blue', marker='s', linestyle='-', label='WMAPE site (Without changes)')
    plt.plot(days[1:], wmape_global_without_changes, color='tab:orange', marker='D', linestyle='--', label='WMAPE global (Without changes)')
    
    plt.xlabel('Days to delivery')
    plt.ylabel('WMAPE')
    plt.title('WMAPE comparison: With vs Without order changes (TS)', pad=20, fontsize=14)
    plt.legend()
    plt.grid(True)
    plt.tight_layout()
    plt.show()

def plot_changed_and_deleted_orders(days, changed_orders_log, deleted_orders_log):
    plt.figure(figsize=(12, 6))
    
    changed_counts = [len(changed_orders_log.get(day, [])) for day in days[1:]]
    deleted_counts = [len(deleted_orders_log.get(day, [])) for day in days[1:]]
    
    width = 0.8
    
    plt.bar(days[1:], changed_counts, width, label='Changed', color='tab:blue')
    plt.bar(days[1:], deleted_counts, width, bottom=changed_counts, label='Deleted', color='tab:orange')
    
    plt.xlabel('Days to delivery')
    plt.ylabel('Number of orders')
    plt.title('Changed and deleted order quantity through days', pad=20, fontsize=14)
    plt.legend()
    plt.grid(True, axis='y', linestyle='--', alpha=0.7)
    plt.tight_layout()
    plt.show()
    
def export_wmape_data_to_excel(filename, days, wmape_site_with_changes, wmape_site_without_changes, wmape_global_with_changes, wmape_global_without_changes):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "WMAPE comparison"

    # Add headers
    headers = ["Days to delivery", "WMAPE site (With changes)", "WMAPE global (With changes)", 
               "WMAPE site (Without changes)", "WMAPE global (Without changes)"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    # Add data
    for row, day in enumerate(days[1:], start=2):
        ws.cell(row=row, column=1, value=day)
        ws.cell(row=row, column=2, value=wmape_site_with_changes[row-2])
        ws.cell(row=row, column=3, value=wmape_global_with_changes[row-2])
        ws.cell(row=row, column=4, value=wmape_site_without_changes[row-2])
        ws.cell(row=row, column=5, value=wmape_global_without_changes[row-2])

    # Adjust column widths
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 25

    wb.save(filename)
    
def export_changed_deleted_orders_to_excel(filename, days, changed_orders_log, deleted_orders_log):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Changed and deleted orders"

    # Add headers
    headers = ["Days to delivery", "Changed orders", "Deleted orders"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    # Add data
    for row, day in enumerate(days[1:], start=2):
        ws.cell(row=row, column=1, value=day)
        ws.cell(row=row, column=2, value=len(changed_orders_log.get(day, [])))
        ws.cell(row=row, column=3, value=len(deleted_orders_log.get(day, [])))

    # Adjust column widths
    for col in range(1, 4):
        ws.column_dimensions[get_column_letter(col)].width = 20

    wb.save(filename)

# Main execution code
start_day = -18
end_day = -3

wmape_site_with_changes, wmape_global_with_changes, _, changed_orders_log, deleted_orders_log = run_allocation_process_over_time(start_day, end_day, total_orders, with_changes=True)
wmape_site_without_changes, wmape_global_without_changes, _, _, _ = run_allocation_process_over_time(start_day, end_day, total_orders, with_changes=False)

days = list(range(start_day, end_day + 1))

# Plot WMAPE comparison
plot_wmape_comparison(days, wmape_site_with_changes, wmape_site_without_changes, wmape_global_with_changes, wmape_global_without_changes)

# Plot changed and deleted orders
plot_changed_and_deleted_orders(days, changed_orders_log, deleted_orders_log)

# Export data to Excel
export_wmape_data_to_excel("WMAPE comparison (TS).xlsx", days, wmape_site_with_changes, wmape_site_without_changes, wmape_global_with_changes, wmape_global_without_changes)
export_changed_deleted_orders_to_excel("Changed and deleted order (TS).xlsx", days, changed_orders_log, deleted_orders_log)