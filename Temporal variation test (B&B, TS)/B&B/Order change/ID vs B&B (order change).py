import random
import matplotlib.pyplot as plt
from pulp import *
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# Set random seed for reproducibility
random.seed(42)

# Define the recipes with their eligibility
recipes_f1_only = list(range(1, 30))
recipes_f1_f2 = list(range(30, 50))
recipes_f2_only = list(range(50, 90))
recipes_f3_only = list(range(90, 101))

# Define total orders and factory capacities
total_orders = 10000
F1_cap = int(0.25 * total_orders)
F2_cap = int(0.5 * total_orders)
F3_cap = float('inf')

factory_capacities = {
    'F1': F1_cap,
    'F2': F2_cap,
    'F3': F3_cap
}

def generate_order_recipes(eligible_recipes, max_recipes=4):
    return random.sample(eligible_recipes, random.randint(1, min(max_recipes, len(eligible_recipes))))

def generate_orders_for_day(real_proportion, simulated_proportion, existing_real_orders=None):
    orders = []
    f1_f3_target = int(0.3 * total_orders)
    f2_f3_target = int(0.6 * total_orders)
    f1_f2_f3_target = int(0.1 * total_orders)
    
    total_real_orders = int(real_proportion * total_orders)
    
    changed_orders = []
    deleted_orders = []
    
    # Create a set of all available IDs
    available_ids = set(range(1, total_orders + 1))
    
    if existing_real_orders:
        # Delete 5% of real orders
        delete_count = int(0.05 * len(existing_real_orders))
        deleted_orders = random.sample(existing_real_orders, delete_count)
        remaining_orders = [order for order in existing_real_orders if order not in deleted_orders]
        
        # Change 30% of remaining real orders
        change_count = int(0.3 * len(remaining_orders))
        changed_orders = random.sample(remaining_orders, change_count)
        
        for order in remaining_orders:
            if order in changed_orders:
                # Change recipe composition and eligibility
                new_recipes = generate_order_recipes(recipes_f1_only + recipes_f1_f2 + recipes_f2_only + recipes_f3_only)
                new_eligibility = ['F3']  # Always eligible for F3
                if any(r in recipes_f1_only + recipes_f1_f2 for r in new_recipes):
                    new_eligibility.append('F1')
                if any(r in recipes_f1_f2 + recipes_f2_only for r in new_recipes):
                    new_eligibility.append('F2')
                order['recipe_ids'] = new_recipes
                order['eligible_factories'] = new_eligibility
            orders.append(order)
            available_ids.remove(order['id'])
    
    # Generate new orders to reach the total_orders
    while len(orders) < total_orders:
        if len([o for o in orders if set(o['eligible_factories']) == {'F1', 'F2', 'F3'}]) < f1_f2_f3_target:
            recipe_ids = generate_order_recipes(recipes_f1_f2)
            eligible_factories = ['F1', 'F2', 'F3']
        elif len([o for o in orders if set(o['eligible_factories']) == {'F1', 'F3'}]) < f1_f3_target - f1_f2_f3_target:
            recipe_ids = generate_order_recipes(recipes_f1_only)
            eligible_factories = ['F1', 'F3']
        elif len([o for o in orders if set(o['eligible_factories']) == {'F2', 'F3'}]) < f2_f3_target - f1_f2_f3_target:
            recipe_ids = generate_order_recipes(recipes_f2_only)
            eligible_factories = ['F2', 'F3']
        else:
            f3_recipe = random.choice(recipes_f3_only)
            all_recipes = recipes_f1_only + recipes_f1_f2 + recipes_f2_only + recipes_f3_only
            remaining_recipes = random.sample(all_recipes, min(3, random.randint(0, 3)))
            recipe_ids = [f3_recipe] + remaining_recipes
            random.shuffle(recipe_ids)
            eligible_factories = ['F3']
        
        new_order = {
            'recipe_ids': recipe_ids,
            'eligible_factories': eligible_factories,
            'is_real': len([o for o in orders if o['is_real']]) < total_real_orders
        }
        
        # Assign a new ID from the available IDs
        new_id = available_ids.pop()
        new_order['id'] = new_id
        
        orders.append(new_order)
    
    # Ensure we have the correct number of real orders
    real_order_count = sum(1 for order in orders if order['is_real'])
    if real_order_count < total_real_orders:
        for order in random.sample([o for o in orders if not o['is_real']], total_real_orders - real_order_count):
            order['is_real'] = True
    elif real_order_count > total_real_orders:
        for order in random.sample([o for o in orders if o['is_real']], real_order_count - total_real_orders):
            order['is_real'] = False
    
    # Shuffle orders to mix real and simulated orders
    random.shuffle(orders)
    
    return orders, changed_orders, deleted_orders

def get_recipe_counts(allocation, by_factory=False):
    if by_factory:
        recipe_counts = {factory: {} for factory in allocation}
        for factory, orders in allocation.items():
            for order in orders:
                for recipe_id in order['recipe_ids']:
                    recipe_counts[factory][recipe_id] = recipe_counts[factory].get(recipe_id, 0) + 1
    else:
        recipe_counts = {}
        for factory, orders in allocation.items():
            for order in orders:
                for recipe_id in order['recipe_ids']:
                    recipe_counts[recipe_id] = recipe_counts.get(recipe_id, 0) + 1
    return recipe_counts

def BB_allocation(orders, factory_capacities, recipe_counts_t_minus_1):
    prob = LpProblem("Order_Allocation", LpMinimize)
    
    x = LpVariable.dicts("allocation", 
                         [(o['id'], f) for o in orders for f in factory_capacities.keys()], 
                         cat='Binary')
    
    for o in orders:
        prob += lpSum(x[o['id'], f] for f in factory_capacities.keys()) == 1
    
    for o in orders:
        for f in factory_capacities.keys():
            if f not in o['eligible_factories']:
                prob += x[o['id'], f] == 0
    
    prob += lpSum(x[o['id'], 'F1'] for o in orders) == factory_capacities['F1']
    prob += lpSum(x[o['id'], 'F2'] for o in orders) == factory_capacities['F2']
    
    recipe_counts_t = {}
    for f in factory_capacities.keys():
        for o in orders:
            for r in o['recipe_ids']:
                if r not in recipe_counts_t:
                    recipe_counts_t[r] = {}
                if f not in recipe_counts_t[r]:
                    recipe_counts_t[r][f] = 0
                recipe_counts_t[r][f] += x[o['id'], f]
    
    abs_diffs = []
    for r in set(recipe_counts_t.keys()) | set(sum((list(counts.keys()) for counts in recipe_counts_t_minus_1.values()), [])):
        for f in factory_capacities.keys():
            t_minus_1_count = recipe_counts_t_minus_1[f].get(r, 0)
            t_count = recipe_counts_t.get(r, {}).get(f, 0)
            
            pos_diff = LpVariable(f"pos_diff_{r}_{f}", lowBound=0)
            neg_diff = LpVariable(f"neg_diff_{r}_{f}", lowBound=0)
            
            prob += t_count - t_minus_1_count == pos_diff - neg_diff
            
            abs_diffs.append(pos_diff + neg_diff)
    
    prob += lpSum(abs_diffs)
    
    prob.solve()
    
    allocation = {f: [] for f in factory_capacities.keys()}
    for o in orders:
        for f in factory_capacities.keys():
            if value(x[o['id'], f]) == 1:
                allocation[f].append(o)
    
    return allocation

def allocate_orders(orders, factory_capacities):
    allocation = {factory: [] for factory in factory_capacities}
    remaining_orders = orders.copy()
    f1_eligible = sorted([order for order in remaining_orders if 'F1' in order['eligible_factories']],
                         key=lambda x: len(x['eligible_factories']))
    allocation['F1'] = [order for order in f1_eligible[:factory_capacities['F1']]]
    remaining_orders = [order for order in remaining_orders if order not in allocation['F1']]
    f2_eligible = [order for order in remaining_orders if 'F2' in order['eligible_factories']]
    allocation['F2'] = [order for order in f2_eligible[:factory_capacities['F2']]]
    remaining_orders = [order for order in remaining_orders if order not in allocation['F2']]
    allocation['F3'] = remaining_orders
    return allocation

def allocate_orders_ID_based(orders, factory_capacities, previous_allocation, changed_orders, deleted_orders):
    allocation = {factory: [] for factory in factory_capacities}
    
    # First pass: Allocate unchanged real orders and eligible changed orders to their previous factories
    for factory, prev_orders in previous_allocation.items():
        for order in prev_orders:
            # Check if the order is real, still exists, and hasn't been deleted
            if order['is_real'] and order['id'] in [o['id'] for o in orders] and order['id'] not in [d['id'] for d in deleted_orders]:
                # Find the current version of the order
                current_order = next(o for o in orders if o['id'] == order['id'])
                
                # Allocate unchanged orders to their previous factory
                if current_order not in changed_orders:
                    allocation[factory].append(current_order)
                # Allocate changed orders to their previous factory if still eligible
                elif factory in current_order['eligible_factories']:
                    allocation[factory].append(current_order)
                # If a changed order is not eligible for its previous factory, it will be handled in the next step with remaining orders
    
    # Identify orders that haven't been allocated yet
    remaining_orders = [order for order in orders if order not in sum(allocation.values(), [])]
    
    # Calculate remaining capacity for each factory
    remaining_capacities = {
        factory: capacity - len(allocation[factory])
        for factory, capacity in factory_capacities.items()
    }
    
    # Allocate remaining orders using the original allocate_orders function
    remaining_allocation = allocate_orders(remaining_orders, remaining_capacities)
    
    # Combine the allocations: previously allocated orders and newly allocated remaining orders
    for factory in allocation:
        allocation[factory].extend(remaining_allocation[factory])
    
    return allocation

def calculate_wmape_site(allocation_t_minus_1, allocation_t):
    recipe_counts_t_minus_1 = get_recipe_counts(allocation_t_minus_1, by_factory=True)
    recipe_counts_t = get_recipe_counts(allocation_t, by_factory=True)
    
    total_abs_diff = 0
    total_items_t = 0
    
    for factory in ['F1', 'F2', 'F3']:
        for recipe_id in set(recipe_counts_t_minus_1[factory].keys()) | set(recipe_counts_t[factory].keys()):
            t_minus_1_count = recipe_counts_t_minus_1[factory].get(recipe_id, 0)
            t_count = recipe_counts_t[factory].get(recipe_id, 0)
            total_abs_diff += abs(t_count - t_minus_1_count)
            total_items_t += t_count
    
    if total_items_t == 0:
        return float('inf')
    else:
        return total_abs_diff / total_items_t

def calculate_wmape_global(allocation_t_minus_1, allocation_t):
    recipe_counts_t_minus_1 = get_recipe_counts(allocation_t_minus_1)
    recipe_counts_t = get_recipe_counts(allocation_t)
    
    all_recipes = set(recipe_counts_t_minus_1.keys()) | set(recipe_counts_t.keys())
    
    total_abs_diff = sum(abs(recipe_counts_t.get(recipe_id, 0) - recipe_counts_t_minus_1.get(recipe_id, 0)) for recipe_id in all_recipes)
    total_items_t = sum(recipe_counts_t.values())
    
    if total_items_t == 0:
        return float('inf')
    else:
        return total_abs_diff / total_items_t
    
def run_allocation_process_over_time(start_day, end_day, total_orders):
    allocations_bb = {}
    allocations_ID_based = {}
    wmape_site_values_bb = []
    wmape_site_values_ID_based = []
    wmape_global_values_bb = []
    real_orders_proportions = []
    previous_real_orders = None
    
    for day in range(start_day, end_day + 1):
        real_proportion = min(1.0, max(0.1, 0.1 + (0.9 / 15) * (18 + day)))
        simulated_proportion = 1 - real_proportion
        
        orders, changed_orders, deleted_orders = generate_orders_for_day(real_proportion, simulated_proportion, previous_real_orders)
        
        allocation_initial = allocate_orders(orders, factory_capacities)
        
        if day == start_day:
            allocation_bb = BB_allocation(orders, factory_capacities, {f: {} for f in factory_capacities})
            allocation_ID_based = allocation_initial
            allocations_bb[day] = allocation_bb
            allocations_ID_based[day] = allocation_ID_based
        else:
            recipe_counts_t_minus_1 = get_recipe_counts(allocations_bb[day-1], by_factory=True)
            allocation_bb = BB_allocation(orders, factory_capacities, recipe_counts_t_minus_1)
            allocation_ID_based = allocate_orders_ID_based(orders, factory_capacities, allocations_ID_based[day-1], changed_orders, deleted_orders)
            
            wmape_site_bb = calculate_wmape_site(allocations_bb[day-1], allocation_bb)
            wmape_site_values_bb.append(wmape_site_bb)
            wmape_site_ID_based = calculate_wmape_site(allocations_ID_based[day-1], allocation_ID_based)
            wmape_site_values_ID_based.append(wmape_site_ID_based)
            
            wmape_global_bb = calculate_wmape_global(allocations_bb[day-1], allocation_bb)
            wmape_global_values_bb.append(wmape_global_bb)
        
        allocations_bb[day] = allocation_bb
        allocations_ID_based[day] = allocation_ID_based
        real_orders_proportions.append(len([o for o in orders if o['is_real']]) / total_orders)
        previous_real_orders = [order for order in orders if order['is_real']]
    
    return wmape_site_values_bb, wmape_site_values_ID_based, wmape_global_values_bb, real_orders_proportions, allocations_bb, allocations_ID_based

def plot_wmape_comparison(days, wmape_site_bb, wmape_site_ID_based, wmape_global_bb):
    plt.figure(figsize=(14, 8))
    
    plt.plot(days[1:], wmape_site_bb, color='tab:red', marker='o', linestyle='-', label='WMAPE site (B&B)')
    plt.plot(days[1:], wmape_site_ID_based, color='tab:blue', marker='s', linestyle='-', label='WMAPE site (ID-based)')
    plt.plot(days[1:], wmape_global_bb, color='tab:green', marker='^', linestyle='--', label='WMAPE global')
    
    plt.xlabel('Days to delivery')
    plt.ylabel('WMAPE')
    plt.title('WMAPE comparison: B&B vs ID-based (With order changes)', pad=20, fontsize=14)
    plt.legend()
    plt.grid(True)
    plt.tight_layout()
    plt.show()
    
def export_wmape_data_to_excel(filename, days, wmape_site_bb, wmape_site_ID_based, wmape_global_bb):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "WMAPE comparison"

    # Add headers
    headers = ["Days to delivery", "WMAPE site (B&B)", "WMAPE site (ID-based)", "WMAPE global"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    # Add data
    for row, day in enumerate(days[1:], start=2):
        ws.cell(row=row, column=1, value=day)
        ws.cell(row=row, column=2, value=wmape_site_bb[row-2])
        ws.cell(row=row, column=3, value=wmape_site_ID_based[row-2])
        ws.cell(row=row, column=4, value=wmape_global_bb[row-2])

    # Adjust column widths
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 20

    wb.save(filename)

# Main execution code
start_day = -18
end_day = -3

wmape_site_bb, wmape_site_ID_based, wmape_global_bb, real_orders_proportions, allocations_bb, allocations_ID_based = run_allocation_process_over_time(start_day, end_day, total_orders)

days = list(range(start_day, end_day + 1))

# Plot WMAPE comparison with ID-based allocation
plot_wmape_comparison(days, wmape_site_bb, wmape_site_ID_based, wmape_global_bb)

# Export WMAPE data to Excel
export_wmape_data_to_excel("ID vs BB (order changes).xlsx", days, wmape_site_bb, wmape_site_ID_based, wmape_global_bb)