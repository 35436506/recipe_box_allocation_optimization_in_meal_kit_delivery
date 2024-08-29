import random
import time
from openpyxl import Workbook
import matplotlib.pyplot as plt
import numpy as np

# Set seed for reproducibility
random.seed(42)

# Define total orders and factory capacities
TOTAL_ORDERS = 10000
F1_CAP = int(0.25 * TOTAL_ORDERS)
F2_CAP = int(0.5 * TOTAL_ORDERS)
FACTORY_CAPACITIES = {'F1': F1_CAP, 'F2': F2_CAP, 'F3': float('inf')}

# Define recipe ranges for each factory
RECIPES_F1_ONLY = list(range(1, 30))
RECIPES_F1_F2 = list(range(30, 50))
RECIPES_F2_ONLY = list(range(50, 90))
RECIPES_F3_ONLY = list(range(90, 101))

# Generate a list of recipe IDs for an order
def generate_order_recipes(eligible_recipes, max_recipes=4):
    return random.sample(eligible_recipes, random.randint(1, min(max_recipes, len(eligible_recipes))))

# Generate orders for a day, including both real and simulated orders
def generate_orders_for_day(real_proportion, simulated_proportion, existing_real_orders=None):
    orders = []
    f1_f3_target = int(0.3 * TOTAL_ORDERS)
    f2_f3_target = int(0.6 * TOTAL_ORDERS)
    f1_f2_f3_target = int(0.1 * TOTAL_ORDERS)
    
    total_real_orders = int(real_proportion * TOTAL_ORDERS)
    available_ids = set(range(1, TOTAL_ORDERS + 1))
    
    if existing_real_orders:
        orders.extend([order.copy() for order in existing_real_orders])
        for order in orders:
            available_ids.remove(order['id'])
    
    while len(orders) < TOTAL_ORDERS:
        if len([o for o in orders if set(o['eligible_factories']) == {'F1', 'F2', 'F3'}]) < f1_f2_f3_target:
            recipe_ids = generate_order_recipes(RECIPES_F1_F2)
            eligible_factories = ['F1', 'F2', 'F3']
        elif len([o for o in orders if set(o['eligible_factories']) == {'F1', 'F3'}]) < f1_f3_target - f1_f2_f3_target:
            recipe_ids = generate_order_recipes(RECIPES_F1_ONLY)
            eligible_factories = ['F1', 'F3']
        elif len([o for o in orders if set(o['eligible_factories']) == {'F2', 'F3'}]) < f2_f3_target - f1_f2_f3_target:
            recipe_ids = generate_order_recipes(RECIPES_F2_ONLY)
            eligible_factories = ['F2', 'F3']
        else:
            f3_recipe = random.choice(RECIPES_F3_ONLY)
            all_recipes = RECIPES_F1_ONLY + RECIPES_F1_F2 + RECIPES_F2_ONLY + RECIPES_F3_ONLY
            remaining_recipes = random.sample(all_recipes, min(3, random.randint(0, 3)))
            recipe_ids = [f3_recipe] + remaining_recipes
            random.shuffle(recipe_ids)
            eligible_factories = ['F3']
        
        new_order = {
            'recipe_ids': recipe_ids,
            'eligible_factories': eligible_factories,
            'is_real': len([o for o in orders if o['is_real']]) < total_real_orders
        }
        
        new_id = available_ids.pop()
        new_order['id'] = new_id
        
        orders.append(new_order)
    
    real_order_count = sum(1 for order in orders if order['is_real'])
    if real_order_count < total_real_orders:
        for order in random.sample([o for o in orders if not o['is_real']], total_real_orders - real_order_count):
            order['is_real'] = True
    elif real_order_count > total_real_orders:
        for order in random.sample([o for o in orders if o['is_real']], real_order_count - total_real_orders):
            order['is_real'] = False
    
    random.shuffle(orders)
    
    return orders

# Print statistics about the generated orders for a given day
def print_order_statistics(orders, day):
    print(f"\nDay {day} eligible orders:")
    f1_f3_eligible = sum(1 for order in orders if set(order['eligible_factories']) == {'F1', 'F3'})
    f2_f3_eligible = sum(1 for order in orders if set(order['eligible_factories']) == {'F2', 'F3'})
    f1_f2_f3_eligible = sum(1 for order in orders if set(order['eligible_factories']) == {'F1', 'F2', 'F3'})
    f3_only_eligible = sum(1 for order in orders if order['eligible_factories'] == ['F3'])
    print(f"Total orders with eligible factories F1, F3: {f1_f3_eligible}")
    print(f"Total orders with eligible factories F2, F3: {f2_f3_eligible}")
    print(f"Total orders with eligible factories F1, F2, F3: {f1_f2_f3_eligible}")
    print(f"Total orders with eligible factories F3: {f3_only_eligible}")
    print(f"Total orders: {len(orders)}")
    print(f"Total real orders: {sum(1 for order in orders if order['is_real'])}")

# Allocate orders to factories based on their eligibility and factory capacities
def allocate_orders(orders, factory_capacities):
    allocation = {factory: [] for factory in factory_capacities}
    remaining_orders = orders.copy()

    f1_eligible = sorted([order for order in remaining_orders if 'F1' in order['eligible_factories']],
                         key=lambda x: len(x['eligible_factories']))
    allocation['F1'] = f1_eligible[:factory_capacities['F1']]
    remaining_orders = [order for order in remaining_orders if order not in allocation['F1']]

    f2_eligible = [order for order in remaining_orders if 'F2' in order['eligible_factories']]
    allocation['F2'] = f2_eligible[:factory_capacities['F2']]
    remaining_orders = [order for order in remaining_orders if order not in allocation['F2']]

    allocation['F3'] = remaining_orders

    return allocation

# Plot the allocation of orders to factories
def plot_allocation(allocation, day):
    plt.figure(figsize=(12, 6))
    factories = ['F1', 'F2', 'F3']
    real_order_counts = [len([order for order in allocation[factory] if order['is_real']]) for factory in factories]
    simulated_order_counts = [len([order for order in allocation[factory] if not order['is_real']]) for factory in factories]

    x = np.arange(len(factories))
    width = 0.35

    plt.bar(x, real_order_counts, width, label='Real orders', color='#1f77b4')
    plt.bar(x, simulated_order_counts, width, bottom=real_order_counts, label='Simulated orders', color='#ff7f0e')

    max_capacity_line = None
    for i, factory in enumerate(factories):
        capacity = FACTORY_CAPACITIES[factory]
        if factory == 'F3':
            capacity = real_order_counts[i] + simulated_order_counts[i] + 3000
        
        line_width = 0.6
        xmin = i - line_width/2
        xmax = i + line_width/2
        
        line = plt.hlines(y=capacity, xmin=xmin, xmax=xmax, 
                    colors='r', linestyles='-', linewidth=2)
        if max_capacity_line is None:
            max_capacity_line = line

    plt.xlabel('Factory')
    plt.ylabel('Order quantity')
    plt.title(f'Allocation for Day {day}')
    plt.xticks(x, factories)
    
    plt.legend([plt.Rectangle((0,0),1,1,fc="#1f77b4"), 
                plt.Rectangle((0,0),1,1,fc="#ff7f0e"),
                max_capacity_line],
               ['Real orders', 'Simulated orders', 'Max capacity'],
               bbox_to_anchor=(1.05, 1), loc='upper left')

    plt.tight_layout()
    plt.show()

# Count the number of recipes allocated to each factory
def get_recipe_counts(allocation, by_factory=False):
    if by_factory:
        recipe_counts = {factory: {} for factory in allocation}
        for factory, orders in allocation.items():
            for order in orders:
                for recipe_id in order['recipe_ids']:
                    recipe_counts[factory][recipe_id] = recipe_counts[factory].get(recipe_id, 0) + 1
    else:
        recipe_counts = {}
        for factory, orders in allocation.items():
            for order in orders:
                for recipe_id in order['recipe_ids']:
                    recipe_counts[recipe_id] = recipe_counts.get(recipe_id, 0) + 1
    return recipe_counts

# Calculate total absolute difference between two recipe counts
def calculate_total_abs_diff(recipe_counts_t_minus_1, recipe_counts_t):
    all_recipes = set(recipe_counts_t_minus_1['F1'].keys()) | set(recipe_counts_t_minus_1['F2'].keys()) | set(recipe_counts_t_minus_1['F3'].keys()) | \
                  set(recipe_counts_t['F1'].keys()) | set(recipe_counts_t['F2'].keys()) | set(recipe_counts_t['F3'].keys())
    
    total_abs_diff = 0
    
    for recipe_id in all_recipes:
        for factory in ['F1', 'F2', 'F3']:
            t_minus_1_count = recipe_counts_t_minus_1[factory].get(recipe_id, 0)
            t_count = recipe_counts_t[factory].get(recipe_id, 0)
            total_abs_diff += abs(t_count - t_minus_1_count)
    
    return total_abs_diff

# Perform Tabu Search to optimize allocation
def tabu_search(allocation_t_minus_1, allocation_t, factory_capacities, max_iterations=500):
    current_allocation = {factory: orders[:] for factory, orders in allocation_t.items()}
    recipe_counts_t_minus_1 = get_recipe_counts(allocation_t_minus_1, by_factory=True)
    current_recipe_counts = get_recipe_counts(current_allocation, by_factory=True)
    current_total_abs_diff = calculate_total_abs_diff(recipe_counts_t_minus_1, current_recipe_counts)
    
    best_allocation = current_allocation.copy()
    best_total_abs_diff = current_total_abs_diff
    
    tabu_list = {}
    tabu_tenure = 20
    
    def swap_orders(factory1, order1, factory2, order2):
        current_allocation[factory1].remove(order1)
        current_allocation[factory2].remove(order2)
        current_allocation[factory1].append(order2)
        current_allocation[factory2].append(order1)
        
        for recipe_id in order1['recipe_ids']:
            current_recipe_counts[factory1][recipe_id] -= 1
            current_recipe_counts[factory2][recipe_id] = current_recipe_counts[factory2].get(recipe_id, 0) + 1
        for recipe_id in order2['recipe_ids']:
            current_recipe_counts[factory2][recipe_id] -= 1
            current_recipe_counts[factory1][recipe_id] = current_recipe_counts[factory1].get(recipe_id, 0) + 1
    
    def calculate_move_impact(order1, factory1, order2, factory2):
        diff = 0
        for recipe_id in order1['recipe_ids']:
            diff -= abs(current_recipe_counts[factory1][recipe_id] - recipe_counts_t_minus_1[factory1].get(recipe_id, 0))
            diff -= abs(current_recipe_counts[factory2].get(recipe_id, 0) - recipe_counts_t_minus_1[factory2].get(recipe_id, 0))
            diff += abs(current_recipe_counts[factory1][recipe_id] - 1 - recipe_counts_t_minus_1[factory1].get(recipe_id, 0))
            diff += abs(current_recipe_counts[factory2].get(recipe_id, 0) + 1 - recipe_counts_t_minus_1[factory2].get(recipe_id, 0))
        for recipe_id in order2['recipe_ids']:
            diff -= abs(current_recipe_counts[factory2][recipe_id] - recipe_counts_t_minus_1[factory2].get(recipe_id, 0))
            diff -= abs(current_recipe_counts[factory1].get(recipe_id, 0) - recipe_counts_t_minus_1[factory1].get(recipe_id, 0))
            diff += abs(current_recipe_counts[factory2][recipe_id] - 1 - recipe_counts_t_minus_1[factory2].get(recipe_id, 0))
            diff += abs(current_recipe_counts[factory1].get(recipe_id, 0) + 1 - recipe_counts_t_minus_1[factory1].get(recipe_id, 0))
        return diff
    
    for iteration in range(max_iterations):       
        best_move = None
        best_move_diff = 0
        
        # Randomly select a subset of orders to consider for swapping
        orders_to_consider = random.sample(sum(current_allocation.values(), []), min(100, len(sum(current_allocation.values(), []))))
        
        for order1 in orders_to_consider:
            factory1 = next(f for f, orders in current_allocation.items() if order1 in orders)
            for factory2 in factory_capacities:
                if factory1 != factory2:
                    for order2 in random.sample(current_allocation[factory2], min(10, len(current_allocation[factory2]))):
                        if factory2 in order1['eligible_factories'] and factory1 in order2['eligible_factories']:
                            move = (order1['id'], factory1, order2['id'], factory2)
                            if move not in tabu_list or current_total_abs_diff + calculate_move_impact(order1, factory1, order2, factory2) < best_total_abs_diff:
                                diff = calculate_move_impact(order1, factory1, order2, factory2)
                                if diff < best_move_diff:
                                    best_move = (order1, factory1, order2, factory2)
                                    best_move_diff = diff
        
        if best_move:
            order1, factory1, order2, factory2 = best_move
            swap_orders(factory1, order1, factory2, order2)
            current_total_abs_diff += best_move_diff
            
            if current_total_abs_diff < best_total_abs_diff:
                best_total_abs_diff = current_total_abs_diff
                best_allocation = {f: orders[:] for f, orders in current_allocation.items()}
            
            tabu_list[(order1['id'], factory1, order2['id'], factory2)] = iteration + tabu_tenure
            tabu_list[(order2['id'], factory2, order1['id'], factory1)] = iteration + tabu_tenure
        
        # Remove expired tabu moves
        tabu_list = {k: v for k, v in tabu_list.items() if v > iteration}
        
        # Diversification strategy
        if iteration % 100 == 0:
            factory1, factory2 = random.sample(list(factory_capacities.keys()), 2)
            order1 = random.choice(current_allocation[factory1])
            order2 = random.choice(current_allocation[factory2])
            if factory2 in order1['eligible_factories'] and factory1 in order2['eligible_factories']:
                swap_orders(factory1, order1, factory2, order2)
                current_total_abs_diff += calculate_move_impact(order1, factory1, order2, factory2)
    
    total_items_t = sum(sum(counts.values()) for counts in get_recipe_counts(best_allocation, by_factory=True).values())
    best_wmape_site = best_total_abs_diff / total_items_t if total_items_t > 0 else float('inf')

    return best_allocation, best_wmape_site

# Calculate WMAPE site and export to Excel
def calculate_wmape_site(allocation_t_minus_1, allocation_t, sheet_wmape):
    recipe_counts_t_minus_1 = get_recipe_counts(allocation_t_minus_1, by_factory=True)
    recipe_counts_t = get_recipe_counts(allocation_t, by_factory=True)
    
    all_recipes = set(recipe_counts_t_minus_1['F1'].keys()) | set(recipe_counts_t_minus_1['F2'].keys()) | set(recipe_counts_t_minus_1['F3'].keys()) | \
                  set(recipe_counts_t['F1'].keys()) | set(recipe_counts_t['F2'].keys()) | set(recipe_counts_t['F3'].keys())
    
    total_abs_diff = 0
    total_items_t = 0
    
    sheet_wmape.append(['Recipe', 'Factory', 'Day t-1', 'Day t', 'Absolute recipe-site difference'])
    
    for recipe_id in sorted(all_recipes):
        for factory in ['F1', 'F2', 'F3']:
            t_minus_1_count = recipe_counts_t_minus_1[factory].get(recipe_id, 0)
            t_count = recipe_counts_t[factory].get(recipe_id, 0)
            abs_diff = abs(t_count - t_minus_1_count)
            total_abs_diff += abs_diff
            total_items_t += t_count
            sheet_wmape.append([recipe_id, factory, t_minus_1_count, t_count, abs_diff])
    
    wmape_site = total_abs_diff / total_items_t if total_items_t > 0 else float('inf')
    
    sheet_wmape.append([''])
    sheet_wmape.append(['', '', 'SUM', total_items_t, total_abs_diff])
    sheet_wmape.append(['WMAPE site before TS', wmape_site])
    sheet_wmape.append([''])
    
    return wmape_site

# Calculate WMAPE global and export to Excel
def calculate_wmape_global(allocation_t_minus_1, allocation_t, sheet_wmape):
    recipe_counts_t_minus_1 = get_recipe_counts(allocation_t_minus_1)
    recipe_counts_t = get_recipe_counts(allocation_t)
    
    all_recipes = set(recipe_counts_t_minus_1.keys()) | set(recipe_counts_t.keys())
    
    total_abs_diff = 0
    total_t_items = sum(recipe_counts_t.values())
    
    sheet_wmape.append(['Recipe', 'Day t-1', 'Day t', 'Absolute recipe difference'])
    
    for recipe_id in sorted(all_recipes):
        t_minus_1_count = recipe_counts_t_minus_1.get(recipe_id, 0)
        t_count = recipe_counts_t.get(recipe_id, 0)
        abs_diff = abs(t_minus_1_count - t_count)
        total_abs_diff += abs_diff
        sheet_wmape.append([recipe_id, t_minus_1_count, t_count, abs_diff])
    
    wmape_global = total_abs_diff / total_t_items if total_t_items > 0 else float('inf')
    
    sheet_wmape.append([''])
    sheet_wmape.append(['', 'SUM', total_t_items, total_abs_diff])
    sheet_wmape.append(['WMAPE global', wmape_global])
    sheet_wmape.append([''])
    
    return wmape_global

# Main execution code
workbook = Workbook()

# Create sheets in the desired order
sheet_day_t_minus_1_orders = workbook.create_sheet("Day t-1 orders", 0)
sheet_day_t_minus_1_allocation = workbook.create_sheet("Day t-1 allocation", 1)
sheet_day_t_orders = workbook.create_sheet("Day t orders", 2)
sheet_day_t_allocation = workbook.create_sheet("Day t allocation (Before)", 3)
sheet_day_t_allocation_ts = workbook.create_sheet("Day t allocation (After)", 4)
sheet_wmape = workbook.create_sheet("WMAPE", 5)

# Generate orders for day -12 (t-1)
orders_t_minus_1 = generate_orders_for_day(0.46, 0.54)  # 46% real orders

# Extract real orders from day t-1
real_orders_t_minus_1 = [order for order in orders_t_minus_1 if order['is_real']]

# Generate orders for day -11 (t), including existing real orders
orders_t = generate_orders_for_day(0.52, 0.48, existing_real_orders=real_orders_t_minus_1)  # 52% real orders

# Print statistics for both days
print_order_statistics(orders_t_minus_1, "-12")
print_order_statistics(orders_t, "-11")

# Perform allocation for day t-1 and day t
allocation_t_minus_1 = allocate_orders(orders_t_minus_1, FACTORY_CAPACITIES)
allocation_t = allocate_orders(orders_t, FACTORY_CAPACITIES)

# Plot allocations
plot_allocation(allocation_t_minus_1, '-12')
plot_allocation(allocation_t, '-11 (Before TS)')

# Calculate WMAPE site and global before TS
wmape_site_before_ts = calculate_wmape_site(allocation_t_minus_1, allocation_t, sheet_wmape)
wmape_global = calculate_wmape_global(allocation_t_minus_1, allocation_t, sheet_wmape)

# Apply Tabu Search to optimize allocation on day t
start_time = time.time()
optimized_allocation_t, optimized_wmape_site = tabu_search(
    allocation_t_minus_1, 
    allocation_t, 
    FACTORY_CAPACITIES, 
    max_iterations=500
)
end_time = time.time()
optimization_time = end_time - start_time

# Export WMAPE site after TS to Excel sheet
sheet_wmape.append([''])
sheet_wmape.append([f"WMAPE site after TS: {optimized_wmape_site:.3f}"])

# Plot solution allocation for day t after TS
plot_allocation(optimized_allocation_t, '-11 (After TS)')

# Populate Day t-1 Orders sheet
sheet_day_t_minus_1_orders.append(["Order ID", "Recipe IDs", "Is real", "Eligible factories"])
for order in orders_t_minus_1:
    sheet_day_t_minus_1_orders.append([order['id'], ", ".join(map(str, order['recipe_ids'])), "Yes" if order['is_real'] else "No", ", ".join(order['eligible_factories'])])

# Populate Day t-1 Allocation sheet
sheet_day_t_minus_1_allocation.append(["Factory", "Allocated orders"])
for factory, orders in allocation_t_minus_1.items():
    order_ids = [order['id'] for order in orders]
    sheet_day_t_minus_1_allocation.append([factory, ", ".join(map(str, order_ids))])

# Populate Day t Orders sheet
sheet_day_t_orders.append(["Order ID", "Recipe IDs", "Is real", "Eligible factories"])
for order in orders_t:
    sheet_day_t_orders.append([order['id'], ", ".join(map(str, order['recipe_ids'])), "Yes" if order['is_real'] else "No", ", ".join(order['eligible_factories'])])

# Populate Day t Allocation (Before) sheet
sheet_day_t_allocation.append(["Factory", "Allocated orders"])
for factory, orders in allocation_t.items():
    order_ids = [order['id'] for order in orders]
    sheet_day_t_allocation.append([factory, ", ".join(map(str, order_ids))])

# Populate Day t Allocation (After) sheet
sheet_day_t_allocation_ts.append(["Factory", "Allocated orders"])
for factory, orders in optimized_allocation_t.items():
    order_ids = [order['id'] for order in orders]
    sheet_day_t_allocation_ts.append([factory, ", ".join(map(str, order_ids))])

# Remove the default sheet if it exists
if "Sheet" in workbook.sheetnames:
    workbook.remove(workbook["Sheet"])

# Save the workbook
workbook.save("TS_allocation_results.xlsx")

# Print final statistics
print("\nDay -12 statistics:")
for factory, orders in allocation_t_minus_1.items():
    print(f"{factory}: {len(orders)} orders, {sum(order['is_real'] for order in orders)} real")

print("\nDay -11 statistics (Before TS):")
for factory, orders in allocation_t.items():
    print(f"{factory}: {len(orders)} orders, {sum(order['is_real'] for order in orders)} real")

print("\nDay -11 statistics (After TS):")
for factory, orders in optimized_allocation_t.items():
    print(f"{factory}: {len(orders)} orders, {sum(order['is_real'] for order in orders)} real")

print("\nWMAPE site improvement:")
print(f"Before TS: {wmape_site_before_ts:.3f}")
print(f"After TS: {optimized_wmape_site:.3f}")
print(f"Improvement: {(wmape_site_before_ts - optimized_wmape_site) / wmape_site_before_ts * 100:.2f}%")
print(f"WMAPE global: {wmape_global:.3f}")
print(f"Optimization time: {optimization_time:.2f} seconds")