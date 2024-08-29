import time as time_module
from pulp import *
import random
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook

# Set random seed for reproducibility
random.seed(42)

# Define total orders and factory capacities
TOTAL_ORDERS = 10000
F1_CAP = int(0.25 * TOTAL_ORDERS)
F2_CAP = int(0.5 * TOTAL_ORDERS)
FACTORY_CAPACITIES = {'F1': F1_CAP, 'F2': F2_CAP, 'F3': float('inf')}

# Define recipe ranges for each factory
RECIPES_F1_ONLY = list(range(1, 30))
RECIPES_F1_F2 = list(range(30, 50))
RECIPES_F2_ONLY = list(range(50, 90))
RECIPES_F3_ONLY = list(range(90, 101))

# Generate a list of recipe IDs for an order
def generate_order_recipes(eligible_recipes, max_recipes=4):
    return random.sample(eligible_recipes, random.randint(1, min(max_recipes, len(eligible_recipes))))

# Generate orders for a day, including both real and simulated orders
def generate_orders_for_day(real_proportion, simulated_proportion, existing_real_orders=None):
    orders = []
    f1_f3_target = int(0.3 * TOTAL_ORDERS)
    f2_f3_target = int(0.6 * TOTAL_ORDERS)
    f1_f2_f3_target = int(0.1 * TOTAL_ORDERS)
    
    total_real_orders = int(real_proportion * TOTAL_ORDERS)
    available_ids = set(range(1, TOTAL_ORDERS + 1))
    
    if existing_real_orders:
        orders.extend([order.copy() for order in existing_real_orders])
        for order in orders:
            available_ids.remove(order['id'])
    
    while len(orders) < TOTAL_ORDERS:
        if len([o for o in orders if set(o['eligible_factories']) == {'F1', 'F2', 'F3'}]) < f1_f2_f3_target:
            recipe_ids = generate_order_recipes(RECIPES_F1_F2)
            eligible_factories = ['F1', 'F2', 'F3']
        elif len([o for o in orders if set(o['eligible_factories']) == {'F1', 'F3'}]) < f1_f3_target - f1_f2_f3_target:
            recipe_ids = generate_order_recipes(RECIPES_F1_ONLY)
            eligible_factories = ['F1', 'F3']
        elif len([o for o in orders if set(o['eligible_factories']) == {'F2', 'F3'}]) < f2_f3_target - f1_f2_f3_target:
            recipe_ids = generate_order_recipes(RECIPES_F2_ONLY)
            eligible_factories = ['F2', 'F3']
        else:
            f3_recipe = random.choice(RECIPES_F3_ONLY)
            all_recipes = RECIPES_F1_ONLY + RECIPES_F1_F2 + RECIPES_F2_ONLY + RECIPES_F3_ONLY
            remaining_recipes = random.sample(all_recipes, min(3, random.randint(0, 3)))
            recipe_ids = [f3_recipe] + remaining_recipes
            random.shuffle(recipe_ids)
            eligible_factories = ['F3']
        
        new_order = {
            'recipe_ids': recipe_ids,
            'eligible_factories': eligible_factories,
            'is_real': len([o for o in orders if o['is_real']]) < total_real_orders
        }
        
        new_id = available_ids.pop()
        new_order['id'] = new_id
        
        orders.append(new_order)
    
    real_order_count = sum(1 for order in orders if order['is_real'])
    if real_order_count < total_real_orders:
        for order in random.sample([o for o in orders if not o['is_real']], total_real_orders - real_order_count):
            order['is_real'] = True
    elif real_order_count > total_real_orders:
        for order in random.sample([o for o in orders if o['is_real']], real_order_count - total_real_orders):
            order['is_real'] = False
    
    random.shuffle(orders)
    
    return orders

# Allocate orders to factories based on their eligibility and factory capacities
def allocate_orders(orders, factory_capacities):
    allocation = {factory: [] for factory in factory_capacities}
    remaining_orders = orders.copy()

    f1_eligible = sorted([order for order in remaining_orders if 'F1' in order['eligible_factories']],
                         key=lambda x: len(x['eligible_factories']))
    allocation['F1'] = f1_eligible[:factory_capacities['F1']]
    remaining_orders = [order for order in remaining_orders if order not in allocation['F1']]

    f2_eligible = [order for order in remaining_orders if 'F2' in order['eligible_factories']]
    allocation['F2'] = f2_eligible[:factory_capacities['F2']]
    remaining_orders = [order for order in remaining_orders if order not in allocation['F2']]

    allocation['F3'] = remaining_orders

    return allocation

# Print statistics about the generated orders for a given day
def print_order_statistics(orders, day):
    print(f"\nDay {day} eligible orders:")
    f1_f3_eligible = sum(1 for order in orders if set(order['eligible_factories']) == {'F1', 'F3'})
    f2_f3_eligible = sum(1 for order in orders if set(order['eligible_factories']) == {'F2', 'F3'})
    f1_f2_f3_eligible = sum(1 for order in orders if set(order['eligible_factories']) == {'F1', 'F2', 'F3'})
    f3_only_eligible = sum(1 for order in orders if order['eligible_factories'] == ['F3'])
    print(f"Total orders with eligible factories F1, F3: {f1_f3_eligible}")
    print(f"Total orders with eligible factories F2, F3: {f2_f3_eligible}")
    print(f"Total orders with eligible factories F1, F2, F3: {f1_f2_f3_eligible}")
    print(f"Total orders with eligible factories F3: {f3_only_eligible}")
    print(f"Total orders: {len(orders)}")
    print(f"Total real orders: {sum(1 for order in orders if order['is_real'])}")

# Count the number of recipes allocated to each factory
def get_recipe_counts(orders, allocation):
    recipe_counts = {f: {} for f in FACTORY_CAPACITIES.keys()}
    for factory, order_ids in allocation.items():
        for order_id in order_ids:
            if isinstance(order_id, dict):
                order = order_id
            else:
                order = next((o for o in orders if o['id'] == order_id), None)
            
            if order is None:
                continue
            
            for recipe in order['recipe_ids']:
                if recipe not in recipe_counts[factory]:
                    recipe_counts[factory][recipe] = 0
                recipe_counts[factory][recipe] += 1
    return recipe_counts

# Calculate the global WMAPE for recipe allocations
def calculate_wmape_global(orders_t_minus_1, allocation_t_minus_1, orders_t, allocation_t):
    total_abs_diff = 0
    total_t_items = 0
    recipe_counts_t_minus_1 = {}
    recipe_counts_t = {}
    
    for factory in allocation_t_minus_1:
        for order in allocation_t_minus_1[factory]:
            for recipe_id in order['recipe_ids']:
                recipe_counts_t_minus_1[recipe_id] = recipe_counts_t_minus_1.get(recipe_id, 0) + 1
    
    for factory in allocation_t:
        for order_id in allocation_t[factory]:
            order = next(o for o in orders_t if o['id'] == order_id)
            for recipe_id in order['recipe_ids']:
                recipe_counts_t[recipe_id] = recipe_counts_t.get(recipe_id, 0) + 1
                total_t_items += 1
    
    all_recipes = set(recipe_counts_t_minus_1.keys()) | set(recipe_counts_t.keys())
    
    for recipe_id in all_recipes:
        t_minus_1_count = recipe_counts_t_minus_1.get(recipe_id, 0)
        t_count = recipe_counts_t.get(recipe_id, 0)
        abs_diff = abs(t_minus_1_count - t_count)
        total_abs_diff += abs_diff
    
    if total_t_items == 0:
        wmape_global = float('inf')
    else:
        wmape_global = total_abs_diff / total_t_items
    
    return wmape_global, total_abs_diff, total_t_items

# Plot the allocation of orders to factories
def plot_allocation(allocation, title, factory_capacities):
    plt.figure(figsize=(12, 6))
    factories = ['F1', 'F2', 'F3']
    real_order_counts = []
    simulated_order_counts = []

    for factory in factories:
        real_count = sum(1 for order in allocation[factory] if order['is_real'])
        simulated_count = sum(1 for order in allocation[factory] if not order['is_real'])
        real_order_counts.append(real_count)
        simulated_order_counts.append(simulated_count)

    x = np.arange(len(factories))
    width = 0.35

    plt.bar(x, real_order_counts, width, label='Real orders', color='#1f77b4')
    plt.bar(x, simulated_order_counts, width, bottom=real_order_counts, label='Simulated orders', color='#ff7f0e')

    max_capacity_line = None
    for i, factory in enumerate(factories):
        capacity = factory_capacities[factory]
        if factory == 'F3':
            capacity = real_order_counts[i] + simulated_order_counts[i] + 3000
        
        line_width = 0.6
        xmin = i - line_width/2
        xmax = i + line_width/2
        
        line = plt.hlines(y=capacity, xmin=xmin, xmax=xmax, 
                    colors='r', linestyles='-', linewidth=2)
        if max_capacity_line is None:
            max_capacity_line = line

    plt.xlabel('Factory')
    plt.ylabel('Order quantity')
    plt.title(title)
    plt.xticks(x, factories)
    
    plt.legend([plt.Rectangle((0,0),1,1,fc="#1f77b4"), 
                plt.Rectangle((0,0),1,1,fc="#ff7f0e"),
                max_capacity_line],
               ['Real orders', 'Simulated orders', 'Max capacity'],
               bbox_to_anchor=(1.05, 1), loc='upper left')

    plt.tight_layout()
    plt.show()
    
# Print detailed statistics about the allocation of orders to factories
def print_allocation_details(allocation):
    print("\nAllocation details for both days:")
    for factory, orders in allocation.items():
        print(f"{factory}: {len(orders)} orders allocated")
        print(f"  F1,F3: {sum(1 for o in orders if set(o['eligible_factories']) == {'F1', 'F3'})}")
        print(f"  F2,F3: {sum(1 for o in orders if set(o['eligible_factories']) == {'F2', 'F3'})}")
        print(f"  F1,F2,F3: {sum(1 for o in orders if set(o['eligible_factories']) == {'F1', 'F2', 'F3'})}")
        print(f"  F3 only: {sum(1 for o in orders if o['eligible_factories'] == ['F3'])}")

# Export the allocation results and WMAPE calculations to an Excel file
def export_to_excel(orders_t_minus_1, orders_t, allocation_t_minus_1, allocation_t, wmape_site, wmape_global):
    workbook = Workbook()

    sheet_t_minus_1_orders = workbook.create_sheet("Day t-1 orders", 0)
    sheet_t_orders = workbook.create_sheet("Day t orders", 1)
    sheet_t_minus_1_allocation = workbook.create_sheet("Day t-1 allocation", 2)
    sheet_t_allocation = workbook.create_sheet("Day t allocation", 3)
    sheet_wmape = workbook.create_sheet("WMAPE", 4)

    sheet_t_minus_1_orders.append(["Order ID", "Recipe IDs", "Is real", "Eligible factories"])
    for order in orders_t_minus_1:
        sheet_t_minus_1_orders.append([order['id'], ", ".join(map(str, order['recipe_ids'])), order['is_real'], ", ".join(order['eligible_factories'])])

    sheet_t_orders.append(["Order ID", "Recipe IDs", "Is real", "Eligible factories"])
    for order in orders_t:
        sheet_t_orders.append([order['id'], ", ".join(map(str, order['recipe_ids'])), order['is_real'], ", ".join(order['eligible_factories'])])

    def write_allocation(sheet, allocation):
        sheet.append(["Factory", "Allocated orders"])
        for factory, orders in allocation.items():
            order_ids = [str(order['id']) if isinstance(order, dict) else str(order) for order in orders]
            sheet.append([factory, ", ".join(order_ids)])

    write_allocation(sheet_t_minus_1_allocation, allocation_t_minus_1)
    write_allocation(sheet_t_allocation, allocation_t)

    sheet_wmape.append(["WMAPE site"])
    sheet_wmape.append(["Recipe", "Factory", "Day t-1", "Day t", "Absolute recipe-site difference"])

    recipe_counts_t_minus_1 = get_recipe_counts(orders_t_minus_1, allocation_t_minus_1)
    recipe_counts_t = get_recipe_counts(orders_t, allocation_t)

    total_abs_diff_site = 0
    total_items_site = 0
    all_recipes = set(sum((list(counts.keys()) for counts in recipe_counts_t_minus_1.values()), [])) | set(sum((list(counts.keys()) for counts in recipe_counts_t.values()), []))

    for r in sorted(all_recipes):
        for f in ['F1', 'F2', 'F3']:
            t_minus_1_count = recipe_counts_t_minus_1[f].get(r, 0)
            t_count = recipe_counts_t[f].get(r, 0)
            abs_diff = abs(t_count - t_minus_1_count)
            sheet_wmape.append([r, f, t_minus_1_count, t_count, abs_diff])
            total_abs_diff_site += abs_diff
            total_items_site += t_count

    sheet_wmape.append([])
    sheet_wmape.append(["Total absolute difference", total_abs_diff_site])
    sheet_wmape.append(["Total items in day t", total_items_site])
    sheet_wmape.append(["WMAPE site", f"{total_abs_diff_site} / {total_items_site} = {wmape_site:.4f}"])

    sheet_wmape.append([])
    sheet_wmape.append(["WMAPE global"])
    
    recipe_counts_t_minus_1_global = {}
    recipe_counts_t_global = {}
    
    for factory in allocation_t_minus_1:
        for order in allocation_t_minus_1[factory]:
            if isinstance(order, dict):
                order_recipes = order['recipe_ids']
            else:
                order = next((o for o in orders_t_minus_1 if o['id'] == order), None)
                if order is None:
                    continue
                order_recipes = order['recipe_ids']
            for recipe_id in order_recipes:
                recipe_counts_t_minus_1_global[recipe_id] = recipe_counts_t_minus_1_global.get(recipe_id, 0) + 1

    total_items_global = 0
    for factory in allocation_t:
        for order_id in allocation_t[factory]:
            order = next((o for o in orders_t if o['id'] == order_id), None)
            if order is None:
                continue
            for recipe_id in order['recipe_ids']:
                recipe_counts_t_global[recipe_id] = recipe_counts_t_global.get(recipe_id, 0) + 1
                total_items_global += 1

    all_recipes_global = set(recipe_counts_t_minus_1_global.keys()) | set(recipe_counts_t_global.keys())
    
    sheet_wmape.append(["Recipe", "Day t-1", "Day t", "Absolute recipe difference"])
    total_abs_diff_global = 0
    for recipe_id in sorted(all_recipes_global):
        t_minus_1_count = recipe_counts_t_minus_1_global.get(recipe_id, 0)
        t_count = recipe_counts_t_global.get(recipe_id, 0)
        abs_diff = abs(t_minus_1_count - t_count)
        sheet_wmape.append([recipe_id, t_minus_1_count, t_count, abs_diff])
        total_abs_diff_global += abs_diff

    sheet_wmape.append([])
    sheet_wmape.append(["Total absolute difference", total_abs_diff_global])
    sheet_wmape.append(["Total items in day t", total_items_global])
    sheet_wmape.append(["WMAPE global", f"{total_abs_diff_global} / {total_items_global} = {wmape_global:.4f}"])

    if "Sheet" in workbook.sheetnames:
        workbook.remove(workbook["Sheet"])

    workbook.save("BB_allocation_results.xlsx")

# Generate orders for day -12
orders_t_minus_1 = generate_orders_for_day(0.46, 0.54)

# Print statistics for day -12
print_order_statistics(orders_t_minus_1, "-12")

# Extract real orders from day -12
real_orders_t_minus_1 = [order for order in orders_t_minus_1 if order['is_real']]

# Generate orders for day -11, including existing real orders
orders_t = generate_orders_for_day(0.52, 0.48, existing_real_orders=real_orders_t_minus_1)

# Print statistics for day -11
print_order_statistics(orders_t, "-11")

# Perform allocation for day -12
allocation_t_minus_1 = allocate_orders(orders_t_minus_1, FACTORY_CAPACITIES)

# Get recipe counts for day -12
recipe_counts_t_minus_1 = get_recipe_counts(orders_t_minus_1, allocation_t_minus_1)

# Start timing for optimization
start_time_optimization = time_module.time()

# Create the optimization model
prob = LpProblem("Order_Allocation", LpMinimize)

# Create binary variables for order allocation
x = LpVariable.dicts("allocation", [(o['id'], f) for o in orders_t for f in FACTORY_CAPACITIES.keys()], cat='Binary')

# Add constraints to ensure each order is allocated to exactly one factory
for o in orders_t:
    prob += lpSum([x[o['id'], f] for f in FACTORY_CAPACITIES.keys()]) == 1

# Add constraints to ensure orders are only allocated to eligible factories
for o in orders_t:
    for f in FACTORY_CAPACITIES.keys():
        if f not in o['eligible_factories']:
            prob += x[o['id'], f] == 0

# Add capacity constraints for F1 and F2
prob += lpSum([x[o['id'], 'F1'] for o in orders_t]) == FACTORY_CAPACITIES['F1']
prob += lpSum([x[o['id'], 'F2'] for o in orders_t]) == FACTORY_CAPACITIES['F2']

# Calculate recipe counts for day t based on allocation
recipe_counts_t = {}
for f in FACTORY_CAPACITIES.keys():
    for o in orders_t:
        for r in o['recipe_ids']:
            if r not in recipe_counts_t:
                recipe_counts_t[r] = {}
            if f not in recipe_counts_t[r]:
                recipe_counts_t[r][f] = 0
            recipe_counts_t[r][f] += x[o['id'], f]

# Calculate absolute differences for WMAPE site
abs_diffs = []
for r in set(recipe_counts_t.keys()) | set(sum((list(counts.keys()) for counts in recipe_counts_t_minus_1.values()), [])):
    for f in FACTORY_CAPACITIES.keys():
        t_minus_1_count = recipe_counts_t_minus_1[f].get(r, 0)
        t_count = recipe_counts_t.get(r, {}).get(f, 0)
        
        pos_diff = LpVariable(f"pos_diff_{r}_{f}", lowBound=0)
        neg_diff = LpVariable(f"neg_diff_{r}_{f}", lowBound=0)
        
        prob += t_count - t_minus_1_count == pos_diff - neg_diff
        
        abs_diffs.append(pos_diff + neg_diff)

# Set the objective to minimize the sum of absolute differences (WMAPE site numerator)
prob += lpSum(abs_diffs)

# Solve the problem
prob.solve()

end_time_optimization = time_module.time()
optimization_time = end_time_optimization - start_time_optimization

# Extract the allocation
allocation_t = {f: [] for f in FACTORY_CAPACITIES.keys()}
for o in orders_t:
    for f in FACTORY_CAPACITIES.keys():
        if value(x[o['id'], f]) == 1:
            allocation_t[f].append(o['id'])

# Calculate and print the WMAPE site
recipe_counts_t = get_recipe_counts(orders_t, allocation_t)
total_abs_diff = 0
total_items = 0

for r in set(sum((list(counts.keys()) for counts in recipe_counts_t_minus_1.values()), [])) | set(sum((list(counts.keys()) for counts in recipe_counts_t.values()), [])):
    for f in FACTORY_CAPACITIES.keys():
        t_minus_1_count = recipe_counts_t_minus_1[f].get(r, 0)
        t_count = recipe_counts_t[f].get(r, 0)
        total_abs_diff += abs(t_count - t_minus_1_count)
        total_items += t_count

wmape_site = total_abs_diff / total_items if total_items > 0 else 0

# Calculate the WMAPE global
wmape_global, global_total_abs_diff, global_total_items = calculate_wmape_global(orders_t_minus_1, allocation_t_minus_1, orders_t, allocation_t)

# Calculate the gap between WMAPE site and WMAPE global
wmape_gap = wmape_site - wmape_global
wmape_gap_percentage = (wmape_gap / wmape_site) * 100 if wmape_site != 0 else 0

# Print statistics
print('-------------------------------------')
print("Day -12 statistics:")
for factory, orders in allocation_t_minus_1.items():
    print(f"{factory}: {len(orders)} orders, {sum(order['is_real'] for order in orders)} real")

print("\nDay -11 statistics (B&B):")
for factory, order_ids in allocation_t.items():
    real_orders = sum(1 for order_id in order_ids if next(o for o in orders_t if o['id'] == order_id)['is_real'])
    print(f"{factory}: {len(order_ids)} orders, {real_orders} real")
    
# Print allocation details
print('-------------------------------------')
print_allocation_details(allocation_t_minus_1)
    
print('-------------------------------------')
print(f"\nWMAPE site: {wmape_site:.3f}")
print(f"WMAPE global: {wmape_global:.3f}")
print(f"Gap between WMAPE site and global: {wmape_gap:.2f} ({wmape_gap_percentage:.1f}%)")
print(f"Optimization time: {optimization_time:.2f} seconds")

# Plot allocations
plot_allocation(allocation_t_minus_1, 'Allocation for Day -12', FACTORY_CAPACITIES)
plot_allocation({f: [o for o in orders_t if o['id'] in ids] for f, ids in allocation_t.items()}, 'Allocation for Day -11 (B&B)', FACTORY_CAPACITIES)

# Export results to Excel
export_to_excel(orders_t_minus_1, orders_t, allocation_t_minus_1, allocation_t, wmape_site, wmape_global)